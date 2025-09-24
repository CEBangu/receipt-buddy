import os
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# This module holds the excel writer class, which is responsible for writing data to the 
# accompanying spreadsheet "receipt-buddy"

class ExcelWriter:
    """
    This class takes in the formatted model output, and writes lines to the excel file.

    This class does NOT create the file, it only writes in available spaces. 
    """ 
    def __init__(self, app_directory):
        
        self.write_path = os.path.join(app_directory, "receipt-buddy.xlsx")
        self.worksheet_name = "Itemized"
        self.table_name = "ReceiptTable"

    def write_rows(self, ModelOutput):
        try:
            workbook = load_workbook(self.write_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"Workbook 'receipt-buddy.xlsx' not found in the directory. Please check if the workbook name has changed")
        try:
            worksheet = workbook[self.worksheet_name]
        except KeyError:
            raise ValueError(f"Worksheet '{self.worksheet_name}' not found in the Excel file. Please check if the worksheet name has changed. It should be 'Itemized'")
        
        try:
            table = worksheet.tables[self.table_name]
        except KeyError:
            raise ValueError(f"Table '{self.table_name}' not found in Excel file. Please check if the table name has changed. It should be 'ReceiptTable'")

        table_params = self._get_table_parameters(table=table, worksheet=worksheet)

        if (table_params.end_row == table_params.first_data_row) and not self._row_has_values(r=table_params.first_data_row, worksheet=worksheet):
            write_row = table_params.first_data_row
        else:
            write_row = table_params.end_row + 1 

        for r in ModelOutput.rows:
            item = worksheet.cell(row=write_row, column=2, value=r[0])

            quantity = worksheet.cell(row=write_row, column=3, value=r[1]) 

            price = worksheet.cell(row=write_row, column=4, value=r[2]) 
            price.number_format = table_params.template_price_format

            price_per_unit = worksheet.cell(row=write_row, column=5, value=r[3]) 
            price_per_unit.number_format = table_params.template_price_per_unit_format

            date = worksheet.cell(row=write_row, column=6, value=r[4]) 
            date.number_format = "DD/MM/YYYY"
            date.alignment = Alignment(horizontal="right")

            write_row += 1
        
        end_row = max(table_params.end_row, write_row - 1) # account for the increment

        table.ref = f"{table_params.start_col}{table_params.start_row}:{table_params.end_col}{end_row}"
        workbook.save(self.write_path)

    def _get_table_parameters(self, table, worksheet):
        start, end = table.ref.split(":")
        
        start_col = ''.join(filter(str.isalpha, start)) 
        start_row = int(''.join(filter(str.isdigit, start)))


        end_col = ''.join(filter(str.isalpha, end))
        end_row = int(''.join(filter(str.isdigit, end)))

        # account for the header
        first_data_row = start_row + 1
        
        tmpl_price_fmt = worksheet.cell(row=first_data_row, column=4).number_format  # D4
        tmpl_ppu_fmt = worksheet.cell(row=first_data_row, column=5).number_format  # E4

        return TableProperties(start_col, start_row, end_col, end_row, first_data_row, tmpl_price_fmt, tmpl_ppu_fmt)
    
    def _row_has_values(self, r, worksheet):
        return any(worksheet.cell(r, c).value not in (None, "") for c in range(2, 7))

@dataclass
class TableProperties:
    """
    This is a dataclass that holds the table properties for the 
    table in the worksheet
    """
    start_col: str
    start_row: int
    end_col: str
    end_row: int
    first_data_row: int
    template_price_format: str
    template_price_per_unit_format: str